import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import subprocess
import os
import filecmp
import openpyxl as xl
from openpyxl.styles import Alignment,Font


app = tk.Tk()

HEIGHT = 500
WIDTH = 600


C = tk.Canvas(app, height=HEIGHT, width=WIDTH)
background_image= tk.PhotoImage(file='./landscape.png')
background_label = tk.Label(app, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

C.pack()

frame = tk.Frame(app,  bg='#42c2f4', bd=5)
frame.place(relx=0.5, rely=0.1, relwidth=0.75, relheight=0.1, anchor='n')



lower_frame = tk.Frame(app, bg='#42c2f4', bd=10)
lower_frame.place(relx=0.5, rely=0.25, relwidth=0.75, relheight=0.6, anchor='n')

label1 = tk.Label(frame,text='Python Code Automation App',font=20)

def get_folder():
    my_folder = filedialog.askdirectory()
    my_folder = str(my_folder)
    print(my_folder)

    def get_folder1():
        my_folder1 = filedialog.askopenfilename()
        print(my_folder1)

        def get_folder2():
            my_folder2 = filedialog.askopenfilename()
            print(my_folder2)

            def get_folder3():
                my_folder3 = filedialog.askopenfilename()
                print(my_folder3)

                # excel part
                wb = xl.load_workbook(my_folder3)
                sheet = wb['Sheet1']
                comp_count = 0
                out_count = 0

                def path_file():
                    # program_file = subprocess.check_output("realpath program_folder", shell=True, text=True)
                    # length3 = len(program_file)
                    # path=program_file[:length3-1]
                    return [os.path.join(my_folder, x) for x in os.listdir(my_folder)]

                def coordinates(val):
                    for row1 in range(5, sheet.max_row + 1):
                        cell = sheet.cell(row1, 2)
                        if cell.value == val:
                            return row1

                lst = path_file()
                for i in lst:
                    print(i)
                    string2 = i[-15:-4]
                    row = coordinates(int(string2))
                    if row > sheet.max_row - 2:
                        break
                    p1 = subprocess.run(f"g++ {i}", shell=True, capture_output=True)
                    data, temp = os.pipe()
                    if (p1.returncode == 0):
                        # print("no compilation error")
                        cell = sheet.cell(row, 3)
                        cell.value = "✔"
                        cell.alignment = Alignment(horizontal='center')

                        with open("output.txt", 'w') as f:
                            # input_file = subprocess.check_output("realpath input.txt", shell=True, text=True)
                            # length2 = len(input_file)
                            file = open(my_folder1, 'r')
                            message = file.read()
                            file.close()
                            os.write(temp, bytes(message, "utf-8"))
                            os.close(temp)
                            process = subprocess.run("./a.out", check=True, stdin=data, shell=True, stdout=f,
                                                     universal_newlines=True)
                            out = subprocess.check_output("realpath output.txt", shell=True, text=True)
                            length = len(out)
                            file1 = out[:length - 1]
                            # output_file = subprocess.check_output("realpath output_original.txt", shell=True, text=True)
                            # length1= len(output_file)
                            # file2 = output_file[:length1-1]
                            file2 = my_folder2
                            comp = filecmp.cmp(file1, file2, shallow=False)
                            if (comp == True):
                                cell = sheet.cell(row, 4)
                                cell.value = "✔"
                                cell.alignment = Alignment(horizontal='center')
                                # print("Program is correct")
                            else:
                                cell = sheet.cell(row, 4)
                                cell.value = "❌"
                                cell.alignment = Alignment(horizontal='center')
                                # print("Program is incorrect")
                                out_count += 1
                    else:
                        cell = sheet.cell(row, 3)
                        cell.value = "❌"
                        cell.alignment = Alignment(horizontal='center')
                        cell = sheet.cell(row, 4)
                        cell.value = "❌"
                        cell.alignment = Alignment(horizontal='center')
                        # print("compilation error")
                        comp_count += 1

                cell = sheet.cell(sheet.max_row, 3)
                cell.value = comp_count
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', size=10.5, bold=True)
                cell = sheet.cell(sheet.max_row, 4)
                cell.value = out_count
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', size=10.5, bold=True)
                wb.save(my_folder3)
                print("Process Completed 🙂️")

            button4 = tk.Button(lower_frame, text='Excel File', font=40, command=get_folder3)
            button4.place(relx=0.26,rely=0.7 ,relheight=0.125, relwidth=0.5)


        button3 = tk.Button(lower_frame, text='Test Cases Output', font=40, command=get_folder2)
        button3.place(relx=0.26,rely=0.5 ,relheight=0.125, relwidth=0.5)

    button2 = tk.Button(lower_frame, text='Test Cases Input', font=40, command=get_folder1)
    button2.place(relx=0.26,rely=0.3 ,relheight=0.125, relwidth=0.5)


button1 = tk.Button(lower_frame, text='Program Folder', font=40, command=get_folder)
button1.place(relx=0.26,rely=0.1 ,relheight=0.125, relwidth=0.5)


app.mainloop()
